from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, Reference, BarChart


def process_wb(filename1):
    wb = load_workbook(filename1)
    ws = wb.active

    # Calculating the average of each module
    rowsmax = ws.max_row
    endrows = rowsmax + 1
    totaldata = rowsmax - 1

    for col in range(3, 7):
        char = get_column_letter(col)
        ws[char + str(endrows)] = f"=SUM({char + '2'}:{char + str(rowsmax)})/{totaldata}"
        ws['B' + str(endrows)] = "Average"
        style = ws['B' + str(endrows)].font = Font(bold=True, color="0099CCFF")

    # Check if a student has passed or failed
    ws['I1'] = "Report"
    for row in range(2, endrows):
        formula = f'SUM({get_column_letter(3)}{row}:{get_column_letter(6)}{row}) / 400 * 100'
        avg_formula = f'IF({formula} >= 50, "PASS", "FAIL")'
        ws[f'I{row}'] = f'={avg_formula}'

    # Count the occurrences of "PASS" and "FAIL"
    count_pass_formula = f'COUNTIF(I2:I{rowsmax}, "PASS")'
    count_fail_formula = f'COUNTIF(I2:I{rowsmax}, "FAIL")'

    ws['j' + str(endrows)] = f'={count_pass_formula}'
    ws['j' + str(endrows + 1)] = f'={count_fail_formula}'
    ws['i' + str(endrows)] = "Total PASS"
    ws['i' + str(endrows)].font = Font(bold=True, color="0099CCFF")
    ws['i' + str(endrows + 1)] = "Total FAIL"
    ws['i' + str(endrows + 1)].font = Font(bold=True, color="0099CCFF")

    # Styling our headings
    colsmax = ws.max_column
    for col in range(1, colsmax):
        char = ws[get_column_letter(col) + '1']. font = Font(bold=True, color="FF0000")

    # Calculating a 10% increase of each bursary funds
    ws["H1"] = "Increase(R)"
    for row in range(2, endrows):
        cell = ws.cell(row, 7)
        corrected_price = f'={cell.value}*0.1+{cell.value}'
        corrected_price_cell = ws.cell(row, 8)
        corrected_price_cell.value = corrected_price

    # Plotting a bar graph of the module averages
    categories = Reference(ws, min_row=endrows+2, min_col=3, max_col=6)
    data = Reference(ws, min_row=endrows, min_col=3, max_col=6)

    chart = BarChart()
    chart.add_data(data)
    chart.style = 10
    chart.set_categories(categories)
    chart.title = 'Modules Percentages'
    chart.y_axis.title = "Percentages"
    chart.x_axis.title = "Modules"
    ws.add_chart(chart, 'c' + str(endrows + 3))

    # Plotting a pie chart for total achievement
    categories2 = Reference(ws, min_row=endrows, max_row=endrows+1, min_col=9)
    data2 = Reference(ws, min_row=endrows, max_row=endrows+1, min_col=10)

    chart1 = PieChart()
    chart1.add_data(data2)
    chart1.set_categories(categories2)
    chart1.style = 10
    chart1.title = 'Pass Rate'
    ws.add_chart(chart1, 'L' + str(endrows+3))

    wb.save("new"+filename1)